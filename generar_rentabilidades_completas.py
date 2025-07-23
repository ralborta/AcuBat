#!/usr/bin/env python3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def generar_rentabilidades_completas():
    print("🔧 GENERANDO REGLAS DE RENTABILIDAD COMPLETAS")
    print("=" * 60)
    
    # Leer el archivo de listado de precios
    archivo_precios = 'Lista Moura 04 (1).xlsx'
    
    try:
        # Leer todas las hojas
        xl = pd.ExcelFile(archivo_precios)
        hojas = xl.sheet_names
        print(f"📊 Hojas disponibles: {hojas}")
        
        # Recolectar todos los productos
        todos_productos = []
        
        for hoja in hojas:
            print(f"\n🔍 Procesando hoja: {hoja}")
            df = pd.read_excel(archivo_precios, sheet_name=hoja)
            
            if len(df) == 0:
                continue
                
            # Contar productos válidos
            productos_validos = 0
            
            for i in range(len(df)):
                try:
                    # Buscar código y precio
                    codigo = None
                    precio = None
                    
                    for col in range(min(5, len(df.columns))):
                        valor = df.iloc[i, col]
                        if pd.notna(valor) and str(valor).strip() and str(valor).strip() != 'nan':
                            if codigo is None:
                                codigo = str(valor).strip()
                            elif precio is None and isinstance(valor, (int, float)) and valor > 0:
                                precio = valor
                                break
                    
                    if codigo and precio:
                        productos_validos += 1
                        todos_productos.append({
                            'codigo': codigo,
                            'precio_base': precio,
                            'hoja_origen': hoja
                        })
                        
                except Exception as e:
                    continue
            
            print(f"  ✅ Productos válidos en {hoja}: {productos_validos}")
        
        print(f"\n🎯 TOTAL DE PRODUCTOS: {len(todos_productos)}")
        
        # Generar reglas de rentabilidad
        print("\n🔧 Generando reglas de rentabilidad...")
        
        # Crear archivo Excel con reglas
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Moura"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        minorista_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        mayorista_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Headers
        headers = [
            "Código", "Precio Base", 
            "Markup Minorista (%)", "Rentabilidad Minorista (%)",
            "Markup Mayorista (%)", "Rentabilidad Mayorista (%)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Datos
        for i, producto in enumerate(todos_productos, 2):
            # Código y precio base
            ws.cell(row=i, column=1, value=producto['codigo'])
            ws.cell(row=i, column=2, value=producto['precio_base'])
            
            # Markups y rentabilidades (valores por defecto)
            markup_minorista = 60.0  # Valor por defecto
            rentabilidad_minorista = 37.5  # Valor por defecto
            markup_mayorista = 22.0  # Valor por defecto
            rentabilidad_mayorista = 18.0  # Valor por defecto
            
            # Ajustar según el tipo de producto
            if producto['codigo'].startswith('M'):
                # Productos Moura - usar valores originales
                markup_minorista = 60.0
                rentabilidad_minorista = 37.5
                markup_mayorista = 22.0
                rentabilidad_mayorista = 18.0
            elif producto['codigo'].startswith('MA'):
                # Productos MA - ajustar valores
                markup_minorista = 55.0
                rentabilidad_minorista = 35.5
                markup_mayorista = 20.0
                rentabilidad_mayorista = 16.7
            elif producto['codigo'].startswith('12MF') or producto['codigo'].startswith('12MS'):
                # Productos 12MF/12MS - ajustar valores
                markup_minorista = 65.0
                rentabilidad_minorista = 39.4
                markup_mayorista = 25.0
                rentabilidad_mayorista = 20.0
            
            # Aplicar estilos
            ws.cell(row=i, column=3, value=markup_minorista).fill = minorista_fill
            ws.cell(row=i, column=4, value=rentabilidad_minorista).fill = minorista_fill
            ws.cell(row=i, column=5, value=markup_mayorista).fill = mayorista_fill
            ws.cell(row=i, column=6, value=rentabilidad_mayorista).fill = mayorista_fill
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"Rentalibilidades_Completas_{timestamp}.xlsx"
        wb.save(nombre_archivo)
        
        print(f"✅ Archivo generado: {nombre_archivo}")
        print(f"📊 Total de productos con reglas: {len(todos_productos)}")
        
        # Mostrar resumen por tipo
        tipos = {}
        for producto in todos_productos:
            tipo = producto['codigo'][:3] if len(producto['codigo']) >= 3 else producto['codigo']
            tipos[tipo] = tipos.get(tipo, 0) + 1
        
        print("\n📋 RESUMEN POR TIPO:")
        for tipo, cantidad in tipos.items():
            print(f"  {tipo}: {cantidad} productos")
        
        return nombre_archivo
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return None

if __name__ == "__main__":
    generar_rentabilidades_completas() 